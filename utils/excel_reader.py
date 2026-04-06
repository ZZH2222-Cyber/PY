# utils/excel_reader.py
"""Excel 测试数据读取：首行为列名，每行一条用例（字典）。"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)


def read_excel_data(
    file_path: str,
    sheet_name: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """读取 .xlsx/.xlsm，将每行转为 dict；跳过全空行。

    参数:
        file_path: Excel 路径（相对项目或绝对路径均可）。
        sheet_name: 工作表名；None 表示第一个工作表。

    返回:
        字典列表，键为第一行表头。

    异常:
        FileNotFoundError: 文件不存在（记录日志后抛出）。
        ValueError: 扩展名不支持、工作表不存在、表头为空或文件损坏。
    """
    logger.info("读取 Excel：%s sheet=%s", file_path, sheet_name)
    path = Path(file_path)

    if not path.is_file():
        logger.error("Excel 文件不存在：%s", file_path)
        raise FileNotFoundError(f"Excel 文件不存在：{file_path}")

    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        logger.error("不支持的格式：%s（需要 .xlsx / .xlsm）", path.suffix)
        raise ValueError(f"不支持的 Excel 格式：{path.suffix}")

    try:
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    except InvalidFileException as exc:
        logger.error("Excel 格式无效：%s — %s", file_path, exc)
        raise ValueError(f"无效的 Excel 文件：{exc}") from exc

    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                logger.error("工作表不存在：%s，可选：%s", sheet_name, workbook.sheetnames)
                raise ValueError(f"工作表不存在：{sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            logger.error("Excel 为空：%s", file_path)
            raise ValueError("Excel 文件为空")

        headers = rows[0]
        if not headers or all(h is None for h in headers):
            logger.error("表头为空：%s", file_path)
            raise ValueError("Excel 表头为空")

        norm_headers: List[str] = []
        for idx, header in enumerate(headers):
            if header is None or (isinstance(header, str) and header.strip() == ""):
                norm_headers.append(f"column_{idx}")
            else:
                norm_headers.append(str(header).strip())

        result: List[Dict[str, Any]] = []
        for row in rows[1:]:
            if row is None:
                continue
            if all(
                cell is None or (isinstance(cell, str) and str(cell).strip() == "")
                for cell in row
            ):
                continue
            row_dict: Dict[str, Any] = {}
            for idx, value in enumerate(row):
                if idx < len(norm_headers):
                    row_dict[norm_headers[idx]] = value
            result.append(row_dict)

        logger.info("Excel 读取完成，共 %d 条数据行", len(result))
        return result
    finally:
        workbook.close()
