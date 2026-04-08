# utils/excel_template.py
"""生成规范化的 Excel 测试用例模板。"""

import logging
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.settings import DATA_DIR

logger = logging.getLogger(__name__)


def create_api_test_template(file_path: Optional[str] = None) -> str:
    """创建 API 接口测试用例模板。

    参数：
        file_path: 输出文件路径，默认为 data/api_test_cases.xlsx。

    返回：
        生成的文件路径。

    异常：
        无。
    """
    if file_path is None:
        file_path = str(Path(DATA_DIR) / "api_test_cases.xlsx")

    # 确保目录存在
    Path(file_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ========== Sheet 1: HTTP方法测试 ==========
    ws_methods = wb.active
    ws_methods.title = "http_methods"

    method_headers = [
        "case_id", "case_name", "priority", "method", "path", 
        "params", "json_body", "expected_status", "assert_type", 
        "assert_field", "assert_value", "enabled", "remark"
    ]

    method_data = [
        ["TC001", "GET请求-基础查询", "高", "GET", "/get", "name=test&value=123", "", 200, "eq", "args.name", "test", "Y", "测试GET请求带查询参数"],
        ["TC002", "GET请求-中文参数", "高", "GET", "/get", "name=张三", "", 200, "eq", "args.name", "张三", "Y", "测试GET请求中文参数"],
        ["TC003", "POST请求-JSON数据", "高", "POST", "/post", "", '{"username":"admin","password":"secret123"}', 200, "eq", "json.username", "admin", "Y", "测试POST请求JSON格式"],
        ["TC004", "POST请求-表单数据", "中", "POST", "/post", "", "", 200, "eq", "form.field1", "value1", "Y", "测试POST请求表单格式,data_body填写field1=value1&field2=value2"],
        ["TC005", "PUT请求-更新数据", "中", "PUT", "/put", "", '{"id":1,"name":"updated"}', 200, "eq", "json.id", "1", "Y", "测试PUT请求"],
        ["TC006", "DELETE请求", "中", "DELETE", "/delete", "", "", 200, "eq", "url", "delete", "Y", "测试DELETE请求"],
    ]

    _write_sheet(ws_methods, method_headers, method_data, "HTTP方法测试用例")

    # ========== Sheet 2: 状态码测试 ==========
    ws_status = wb.create_sheet("status_codes")

    status_data = [
        ["TC101", "状态码200-成功", "高", "GET", "/status/200", "", "", 200, "eq", "status_code", "200", "Y", "测试成功状态码"],
        ["TC102", "状态码301-重定向", "中", "GET", "/status/301", "", "", 301, "eq", "status_code", "301", "Y", "测试重定向状态码"],
        ["TC103", "状态码404-未找到", "中", "GET", "/status/404", "", "", 404, "eq", "status_code", "404", "Y", "测试资源不存在"],
        ["TC104", "状态码500-服务器错误", "中", "GET", "/status/500", "", "", 500, "eq", "status_code", "500", "Y", "测试服务器错误"],
    ]

    _write_sheet(ws_status, method_headers, status_data, "状态码测试用例")

    # ========== Sheet 3: 请求头测试 ==========
    ws_headers = wb.create_sheet("headers")

    headers_data = [
        ["TC201", "自定义请求头", "高", "GET", "/headers", "", "", 200, "contains", "headers.X-Custom", "custom", "Y", "测试自定义请求头,headers填写"],
        ["TC202", "User-Agent验证", "中", "GET", "/user-agent", "", "", 200, "contains", "user-agent", "", "Y", "验证User-Agent响应"],
        ["TC203", "IP地址查询", "低", "GET", "/ip", "", "", 200, "contains", "origin", ".", "Y", "获取客户端IP"],
    ]

    _write_sheet(ws_headers, method_headers, headers_data, "请求头测试用例")

    # ========== Sheet 4: Cookie测试 ==========
    ws_cookies = wb.create_sheet("cookies")

    cookies_data = [
        ["TC301", "设置Cookie", "中", "GET", "/cookies/set", "session_id=abc123", "", 200, "eq", "cookies.session_id", "abc123", "Y", "测试设置Cookie"],
        ["TC302", "获取Cookie", "中", "GET", "/cookies", "", "", 200, "contains", "cookies", "", "Y", "测试获取Cookie"],
    ]

    _write_sheet(ws_cookies, method_headers, cookies_data, "Cookie测试用例")

    # ========== Sheet 5: 字段说明 ==========
    ws_doc = wb.create_sheet("字段说明")

    doc_headers = ["字段名", "必填", "类型", "说明", "示例"]
    doc_data = [
        ["case_id", "是", "字符串", "用例唯一标识", "TC001"],
        ["case_name", "是", "字符串", "用例名称（中文描述）", "GET请求-基础查询"],
        ["priority", "是", "字符串", "优先级：高/中/低", "高"],
        ["method", "是", "字符串", "HTTP方法：GET/POST/PUT/DELETE", "GET"],
        ["path", "是", "字符串", "接口路径（相对路径）", "/get"],
        ["params", "否", "字符串", "查询参数（&分隔）", "page=1&size=10"],
        ["json_body", "否", "JSON字符串", "JSON请求体", '{"name": "test"}'],
        ["expected_status", "是", "整数", "期望的HTTP状态码", "200"],
        ["assert_type", "是", "字符串", "断言类型：eq/contains/regex", "eq"],
        ["assert_field", "是", "字符串", "断言字段（支持嵌套）", "args.name"],
        ["assert_value", "是", "字符串", "期望的断言值", "test"],
        ["enabled", "是", "字符串", "是否启用：Y/N", "Y"],
        ["remark", "否", "字符串", "备注说明", "测试GET请求"],
    ]

    _write_sheet(ws_doc, doc_headers, doc_data, "Excel 测试用例字段说明")

    # 保存文件
    wb.save(file_path)
    logger.info("Excel 测试用例模板已生成：%s", file_path)

    return file_path


def _write_sheet(ws, headers: List[str], data: List[List], title: str) -> None:
    """写入工作表数据并设置样式。

    参数：
        ws: openpyxl 工作表对象。
        headers: 表头列表。
        data: 数据行列表。
        title: 工作表标题。

    返回：
        无
    """
    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 设置列宽
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in data:
            if col_idx <= len(row):
                cell_length = len(str(row[col_idx - 1]))
                max_length = max(max_length, cell_length)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 40)

    # 冻结首行
    ws.freeze_panes = "A2"


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    create_api_test_template()