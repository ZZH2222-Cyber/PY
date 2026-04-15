"""AI 用例生成 → 写入规范化 Excel（供 DDT 执行）。

用法（示例）：
python -m utils.ai_to_excel --method POST --path /demo --description "..."

默认写入 data/api_test_cases.xlsx 的 api 工作表：
- 若不存在则创建
- 写入前会删除同一 method+path 的旧行，再追加本次生成的用例（避免重复堆积）
"""

from __future__ import annotations

import argparse
import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from config.settings import DATA_DIR
from utils.ai_helper import generate_test_cases_by_ai

logger = logging.getLogger(__name__)


TEMPLATE_HEADERS: List[str] = [
    "case_id",
    "case_name",
    "priority",
    "method",
    "path",
    "params",
    "headers",
    "json_body",
    "expected_status",
    "expected_code",
    "assert_type",
    "assert_field",
    "assert_value",
    "extract_var",
    "extract_field",
    "enabled",
    "remark",
]


def _to_json_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _api_key(method: str, path: str) -> Tuple[str, str]:
    return method.strip().upper(), (path.strip() or "/")


def _ensure_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(TEMPLATE_HEADERS)
    # 若表头为空或不完整，尽量修复到模板表头（不覆盖已有数据列，只补缺）
    if ws.max_row < 1:
        ws.append(TEMPLATE_HEADERS)
    else:
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        norm = [str(h).strip() if h is not None else "" for h in headers]
        if not norm or all(h == "" for h in norm):
            ws.delete_rows(1, ws.max_row)
            ws.append(TEMPLATE_HEADERS)
        else:
            # 若缺少关键列，重写表头为模板（保守策略：只在列数不足时）
            if len(norm) < len(TEMPLATE_HEADERS):
                ws.delete_rows(1)
                ws.insert_rows(1)
                for idx, h in enumerate(TEMPLATE_HEADERS, 1):
                    ws.cell(row=1, column=idx, value=h)
    return ws


def _delete_rows_for_api(ws, method: str, path: str) -> int:
    """删除同一 method+path 的旧用例行（从第2行开始）。"""
    deleted = 0
    m, p = _api_key(method, path)
    # 从底往上删，避免行号变化
    for r in range(ws.max_row, 1, -1):
        cell_method = ws.cell(row=r, column=TEMPLATE_HEADERS.index("method") + 1).value
        cell_path = ws.cell(row=r, column=TEMPLATE_HEADERS.index("path") + 1).value
        if str(cell_method or "").strip().upper() == m and str(cell_path or "").strip() == p:
            ws.delete_rows(r, 1)
            deleted += 1
    return deleted


def _map_ai_case_to_row(ai_case: Dict[str, Any], index: int, method: str, path: str) -> Dict[str, Any]:
    # 兼容 AI 输出字段：title/body/headers/expected_status/category
    title = str(ai_case.get("title") or ai_case.get("case_name") or f"AI用例-{index}").strip()
    headers = ai_case.get("headers") or {}
    body = ai_case.get("body")
    expected_status = ai_case.get("expected_status", 200)
    category = str(ai_case.get("category") or "").strip()

    m, p = _api_key(ai_case.get("method") or method, ai_case.get("path") or path)
    case_id = f"AI_{m}_{p.strip('/').replace('/', '_')}_{index:03d}" if p != "/" else f"AI_{m}_root_{index:03d}"

    remark = "AI生成"
    if category:
        remark = f"{remark} | {category}"

    row: Dict[str, Any] = {
        "case_id": case_id,
        "case_name": title,
        "priority": "中",
        "method": m,
        "path": p,
        "params": "",
        "headers": _to_json_str(headers) if headers else "",
        "json_body": _to_json_str(body) if body is not None else "",
        "expected_status": int(expected_status) if str(expected_status).isdigit() else 200,
        "expected_code": 0,
        "assert_type": "eq",
        "assert_field": "",
        "assert_value": "",
        "extract_var": "",
        "extract_field": "",
        "enabled": "Y",
        "remark": remark,
    }
    return row


def write_ai_cases_to_api_sheet(
    *,
    excel_path: str,
    sheet_name: str,
    method: str,
    path: str,
    description: str,
    name: str = "ai_generated",
    delete_only: bool = False,
    trust_expected_status: bool = False,
) -> str:
    """调用 AI 生成用例并写入规范化 Excel（api sheet）。"""
    cases: List[Dict[str, Any]] = []
    if not delete_only:
        api_info = {"name": name, "method": method, "path": path, "description": description}
        cases = generate_test_cases_by_ai(api_info)
        if not cases:
            raise ValueError("AI 未生成任何用例")

    xlsx = Path(excel_path)
    xlsx.parent.mkdir(parents=True, exist_ok=True)

    if xlsx.exists():
        wb = load_workbook(str(xlsx))
    else:
        # 没有模板文件也可创建最小可用文件（只建 api sheet）
        from openpyxl import Workbook

        wb = Workbook()
        # 默认 sheet 改名为 api，避免多余 sheet
        wb.active.title = sheet_name

    ws = _ensure_sheet(wb, sheet_name)

    deleted = _delete_rows_for_api(ws, method, path)
    logger.info("已删除旧用例行：%d（method=%s path=%s）", deleted, method, path)

    start_row = ws.max_row + 1
    if not delete_only:
        for i, c in enumerate(cases, 1):
            mapped = _map_ai_case_to_row(c, i, method, path)
            if not trust_expected_status:
                mapped["expected_status"] = 200
            ws.append([mapped.get(h, "") for h in TEMPLATE_HEADERS])

    wb.save(str(xlsx))
    wb.close()
    if delete_only:
        logger.info("仅删除完成：%s sheet=%s deleted=%d", xlsx, sheet_name, deleted)
    else:
        logger.info("已写入 %d 条 AI 用例到 %s sheet=%s 起始行=%d", len(cases), xlsx, sheet_name, start_row)
    return str(xlsx)


def main() -> int:
    parser = argparse.ArgumentParser(description="调用 DeepSeek 生成用例并写入 data/api_test_cases.xlsx 的 api sheet")
    parser.add_argument("--method", required=True, help="HTTP 方法，例如 GET/POST")
    parser.add_argument("--path", required=True, help="接口路径，例如 /login")
    parser.add_argument("--description", required=True, help="接口描述（用于 AI 生成）")
    parser.add_argument("--name", default="ai_generated", help="接口名称（可选）")
    parser.add_argument("--excel", default=str(Path(DATA_DIR) / "api_test_cases.xlsx"), help="输出 Excel 路径")
    parser.add_argument("--sheet", default="api", help="写入的工作表名，默认 api")
    parser.add_argument("--delete-only", action="store_true", help="只删除指定 method+path 的旧行，不生成/写入新用例")
    parser.add_argument(
        "--trust-expected-status",
        action="store_true",
        help="信任 AI 生成的 expected_status（默认关闭：统一写 200，便于先跑通执行链路）",
    )
    args = parser.parse_args()

    write_ai_cases_to_api_sheet(
        excel_path=args.excel,
        sheet_name=args.sheet,
        method=args.method,
        path=args.path,
        description=args.description,
        name=args.name,
        delete_only=args.delete_only,
        trust_expected_status=args.trust_expected_status,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

