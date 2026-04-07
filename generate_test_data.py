import openpyxl
import random
import string

# 生成随机字符串
def generate_random_string(length=10):
    return ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(length))

# 生成随机邮箱
def generate_random_email():
    return f"{generate_random_string(8)}@{generate_random_string(5)}.com"

# 生成随机手机号
def generate_random_phone():
    return f"138{random.randint(10000000, 99999999)}"

# 生成测试数据并写入Excel文件
def generate_test_data():
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    
    # 用户信息测试数据
    ws_users = wb.active
    ws_users.title = "Users"
    ws_users.append(['ID', 'Name', 'Age', 'Email', 'Phone', 'Password'])
    
    for i in range(1, 11):
        ws_users.append([
            i,
            f"User{i}",
            random.randint(18, 60),
            generate_random_email(),
            generate_random_phone(),
            f"Password{i}"
        ])
    
    # 产品信息测试数据
    ws_products = wb.create_sheet(title="Products")
    ws_products.append(['ID', 'Name', 'Price', 'Stock', 'Description'])
    
    products = [
        ("Laptop", 5999.99, 50, "High performance laptop"),
        ("Smartphone", 3999.99, 100, "Latest smartphone"),
        ("Tablet", 2999.99, 80, "Portable tablet"),
        ("Smartwatch", 1999.99, 120, "Fitness smartwatch"),
        ("Headphones", 999.99, 200, "Wireless headphones")
    ]
    
    for i, (name, price, stock, description) in enumerate(products, 1):
        ws_products.append([i, name, price, stock, description])
    
    # 订单测试数据
    ws_orders = wb.create_sheet(title="Orders")
    ws_orders.append(['OrderID', 'UserID', 'ProductID', 'Quantity', 'TotalPrice', 'Status'])
    
    statuses = ["Pending", "Processing", "Shipped", "Delivered", "Cancelled"]
    
    for i in range(1, 11):
        user_id = random.randint(1, 10)
        product_id = random.randint(1, 5)
        quantity = random.randint(1, 5)
        total_price = round(random.uniform(999.99, 19999.99), 2)
        status = random.choice(statuses)
        
        ws_orders.append([i, user_id, product_id, quantity, total_price, status])
    
    # 登录测试数据
    ws_login = wb.create_sheet(title="Login")
    ws_login.append(['TestCase', 'Username', 'Password', 'ExpectedResult'])
    
    login_tests = [
        ("Valid login", "admin", "password123", "Success"),
        ("Invalid password", "admin", "wrongpassword", "Failure"),
        ("Invalid username", "wronguser", "password123", "Failure"),
        ("Empty username", "", "password123", "Failure"),
        ("Empty password", "admin", "", "Failure"),
        ("Empty both", "", "", "Failure")
    ]
    
    for test_case, username, password, expected in login_tests:
        ws_login.append([test_case, username, password, expected])
    
    # 边界值测试数据
    ws_boundary = wb.create_sheet(title="Boundary")
    ws_boundary.append(['TestCase', 'Input', 'ExpectedResult'])
    
    boundary_tests = [
        ("Minimum age", 18, "Valid"),
        ("Below minimum age", 17, "Invalid"),
        ("Maximum age", 60, "Valid"),
        ("Above maximum age", 61, "Invalid"),
        ("Minimum password length", "Pass1", "Invalid"),
        ("Valid password length", "Password123", "Valid"),
        ("Maximum password length", "Password12345678901234567890", "Valid"),
        ("Above maximum password length", "Password123456789012345678901", "Invalid")
    ]
    
    for test_case, input_val, expected in boundary_tests:
        ws_boundary.append([test_case, input_val, expected])
    
    # 异常场景测试数据
    ws_exception = wb.create_sheet(title="Exception")
    ws_exception.append(['TestCase', 'Input', 'ExpectedResult'])
    
    exception_tests = [
        ("SQL injection", "1' OR 1=1", "Security error"),
        ("XSS attack", "<script>alert('XSS')</script>", "Security error"),
        ("Prompt injection", "Ignore previous instructions", "Security error"),
        ("Large input", "A" * 1000, "Validation error"),
        ("Special characters", "!@#$%^&*()", "Validation error")
    ]
    
    for test_case, input_val, expected in exception_tests:
        ws_exception.append([test_case, input_val, expected])
    
    # 保存文件
    wb.save('data/test_data.xlsx')
    print("测试数据已生成并保存到 data/test_data.xlsx")

if __name__ == "__main__":
    generate_test_data()
