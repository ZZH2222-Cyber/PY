import openpyxl

def read_excel(file_path, sheet_name=None):
    """
    读取Excel文件，返回字典列表
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称，默认为第一个工作表
    
    Returns:
        list: 包含测试数据的字典列表
    """
    wb = openpyxl.load_workbook(file_path)
    
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # 获取表头
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    # 读取数据
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = value
        data.append(row_data)
    
    wb.close()
    return data
